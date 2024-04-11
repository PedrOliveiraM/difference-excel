import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
#
#/home/pedro/Documents/Estagio/cmpPlanilhas/dataFrame base.xlsx
#/home/pedro/Documents/Estagio/cmpPlanilhas/dataFrame novo.xlsx

def abrir_arquivo_base():
    filename = filedialog.askopenfilename(initialdir="/", title="Selecione o arquivo Excel",
                                          filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
    if filename:
        caminho_arquivo_base.delete(0, tk.END)
        caminho_arquivo_base.insert(0, filename)
        wb = load_workbook(filename)
        tabelas = wb.sheetnames
        return tabelas

def selecionarTabelas():
    tabelas = abrir_arquivo_base()
    if tabelas:
        combo['values'] = tabelas
        
        
def abrir_arquivo_cmp():
    filename = filedialog.askopenfilename(initialdir="/", title="Selecione o arquivo Excel",
                                          filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
    if filename:
        caminho_arquivo_cmp.delete(0, tk.END)
        caminho_arquivo_cmp.insert(0, filename)
        
def local_save():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
    if file_path:
        caminho_arquivo_save.delete(0, tk.END)
        caminho_arquivo_save.insert(0, file_path)
        
def comparar_planilhas():
    arquivo_base = caminho_arquivo_base.get()
    arquivo_cmp = caminho_arquivo_cmp.get()
    arquivo_save = caminho_arquivo_save.get()
    nome_tabela_base = nome_tabela_base_entry.get()
    nome_tabela_cmp = nome_tabela_cmp_entry.get()

    if arquivo_base and arquivo_cmp and nome_tabela_base and nome_tabela_cmp:
        try:
            wb_base = load_workbook(filename=arquivo_base)
            wb_cmp = load_workbook(filename=arquivo_cmp)

            planilha_base = wb_base[nome_tabela_base]
            planilha_cmp = wb_cmp[nome_tabela_cmp]

            for row_base, row_cmp in zip(planilha_base.iter_rows(), planilha_cmp.iter_rows()):
                for cell_base, cell_cmp in zip(row_base, row_cmp):
                    if cell_base.value != cell_cmp.value:
                        cell_cmp.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            wb_cmp.save(filename=arquivo_save)

            tk.messagebox.showinfo("Comparação Concluída !", "As diferenças foram destacadas e salvas")
        
        except Exception as e:
            tk.messagebox.showerror("Erro", f"Ocorreu um erro ao comparar as planilhas: {str(e)}")
    else:
        tk.messagebox.showwarning("Aviso", "Por favor, preencha todos os campos.")

root = tk.Tk() # instanciando minha janela
root.minsize(600, 400)  # Define o tamanho mínimo da janela como 400x300 pixels
root.title("Manipulador de Arquivo Excel")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)


title = tk.Label(frame, text= "Comparador de Planilhas Excel",font=("Arial",28,"bold"))
title.grid(row=0, column=0, columnspan=2, padx=10, pady=5)

# ---------------------------------------------------------------------------------------------------------
                    # BASE
                    
                    
label_base = tk.Label(frame, text="Nome da Tabela na Planilha Base:")
label_base.grid(row=2, column=0, padx=10, pady=5)  # Adicionando padx e pady

comboBase = ttk.Combobox(frame, state="readonly", width=38)
comboBase.grid(row=2,column=1,padx=5, pady=5)


botao_arq_base = tk.Button(frame, text="Selecionar Arquivo Base Excel", command=selecionarTabelas)
botao_arq_base.grid(row=1, column=0, padx=10, pady=5)  # Adicionando padx e pady

caminho_arquivo_base = tk.Entry(frame, width=40)
caminho_arquivo_base.grid(row=1, column=1, padx=10, pady=5)  # Adicionando padx e pady

# ---------------------------------------------------------------------------------------------------------

                    #COMPARAÇÃO
                            
 
botao_arq_comp = tk.Button(frame, text="Selecionar Arquivo para Comparar", command=abrir_arquivo_cmp)
botao_arq_comp.grid(row=3, column=0, padx=10, pady=5)  # Adicionando padx e pady

caminho_arquivo_cmp = tk.Entry(frame, width=40)
caminho_arquivo_cmp.grid(row=3, column=1, padx=10, pady=5)  # Adicionando padx e pady
                   
label_cmp = tk.Label(frame, text="Nome da Tabela na Planilha a Comparar:")
label_cmp.grid(row=4, column=0, padx=10, pady=5)  # Adicionando padx e pady

comboCMP = ttk.Combobox(frame, state="readonly", width=38)
comboCMP.grid(row=4,column=1,padx=5, pady=5)



# ---------------------------------------------------------------------------------------------------------

                # SAVE

botao_arq_save = tk.Button(frame, text="Selecionar Local Onde Salvar o Arquivo", command=local_save) 
botao_arq_save.grid(row=5, column=0, padx=10, pady=5)

caminho_arquivo_save= tk.Entry(frame, width=40)
caminho_arquivo_save.grid(row=5, column=1, padx=10, pady=5)  # Adicionando padx e pady

# ---------------------------------------------------------------------------------------------------------

botao_comparar = tk.Button(frame, text="Comparar", command=comparar_planilhas , width=20, height=2 , font=("Arial",12,"bold"))
botao_comparar.grid(row=6, column=0, columnspan=2, padx=10, pady=5)  # Adicionando padx e pady

root.mainloop()
